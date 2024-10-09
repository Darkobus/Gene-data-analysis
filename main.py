import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import filedialog
import threading

def browse_file(entry):
    filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    entry.delete(0, tk.END)
    entry.insert(0, filename)


def browse_directory(entry):
    directory = filedialog.askdirectory()
    entry.delete(0, tk.END)
    entry.insert(0, directory)


def start_process():
    log_text.insert(tk.END, "Working...\n")
    log_text.update()

    df2 = pd.read_excel(entry_file2.get())
    log_text.insert(tk.END, "Loaded Compare.xlsx.\n")
    log_text.update()

    columns_input = entry_columns.get().split(';')
    columns_input = [col.strip() + "_log2FoldChange" for col in columns_input]

    df1 = pd.read_excel(entry_file1.get(), usecols=["gene_id", "go_term"])
    log_text.insert(tk.END, "Loaded GO.xlsx.\n")
    log_text.update()

    fraza = entry_function.get()

    df1['go_term'] = df1['go_term'].astype(str)
    wynik = df1[df1['go_term'].str.contains(fraza, na=False, case=False)]
    ilosc_fraz = len(wynik)

    log_text.insert(tk.END, f"Found {ilosc_fraz} matching phrases for '{fraza}'.\n")
    log_text.update()

    pasujace_gene_id = wynik['gene_id']
    filtrowane_df2 = df2[df2['gene_id'].isin(pasujace_gene_id)]
    log_text.insert(tk.END, "Found compare gene_id in Compare.xlsx.\n")
    log_text.update()

    output_file = entry_dir.get() + '/result.xlsx'
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for col in columns_input:
            col_1 = col_2 = col_pvalue = col_padj = None

            if col in df2.columns:
                col_index = df2.columns.get_loc(col)

                if col_index > 1:
                    col_1 = df2.columns[col_index - 2]
                if col_index > 0:
                    col_2 = df2.columns[col_index - 1]
                if col_index + 1 < len(df2.columns):
                    col_pvalue = df2.columns[col_index + 1]
                if col_index + 2 < len(df2.columns):
                    col_padj = df2.columns[col_index + 2]

            df2_filtrowane = pd.merge(filtrowane_df2, wynik[['gene_id', 'go_term']], on='gene_id', how='left')

            if col_padj and col_padj in df2_filtrowane.columns:
                df2_filtrowane = df2_filtrowane[df2_filtrowane[col_padj] < 0.05]
                log_text.insert(tk.END, "Used filtration with padj < 0.05.\n")
                log_text.update()

            columns_to_include = [
                'gene_id', col_1, col_2, col, col_pvalue, col_padj,
                'gene_name', 'gene_chr', 'gene_start', 'gene_end',
                'gene_strand', 'gene_length', 'gene_biotype',
                'gene_description', 'Family', 'go_term'
            ]
            columns_to_include = [c for c in columns_to_include if c is not None]

            df2_filtrowane = df2_filtrowane[columns_to_include]
            log_text.insert(tk.END, f"Added column {col} to result.\n")
            log_text.update()

            df2_filtrowane.to_excel(writer, index=False, sheet_name=col[:-len('_log2FoldChange')])
            workbook = writer.book
            worksheet = writer.sheets[col[:-len('_log2FoldChange')]]

            for row in range(2, len(df2_filtrowane) + 2):
                cell_value = df2_filtrowane.iloc[row - 2][col]
                cell = worksheet.cell(row=row, column=columns_to_include.index(col) + 1)

                if cell_value > 0:
                    cell.fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
                elif cell_value < 0:
                    cell.fill = PatternFill(start_color='FF6347', end_color='FF6347', fill_type='solid')

            for col_num in range(1, len(columns_to_include) + 1):
                for row in range(1, len(df2_filtrowane) + 2):
                    worksheet.cell(row=row, column=col_num).alignment = Alignment(wrap_text=True)

            column_widths = {
                'gene_id': 20,
                col_1: 23,
                col_2: 25,
                col: 33,
                col_pvalue: 25,
                col_padj: 25,
                'gene_name': 15,
                'gene_chr': 10,
                'gene_start': 15,
                'gene_end': 10,
                'gene_strand': 15,
                'gene_length': 15,
                'gene_biotype': 20,
                'gene_description': 30,
                'Family': 10,
                'go_term': 30
            }

            for col_name, width in column_widths.items():
                if col_name in df2_filtrowane.columns:
                    col_index = df2_filtrowane.columns.get_loc(col_name) + 1
                    worksheet.column_dimensions[get_column_letter(col_index)].width = width

            header_fill = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid')
            for col_num in range(1, len(columns_to_include) + 1):
                header_cell = worksheet.cell(row=1, column=col_num)
                header_cell.fill = header_fill

            thin_border = Border(bottom=Side(style='thin', color='C2B280'))

            for col_num in range(1, len(columns_to_include) + 1):
                worksheet.cell(row=1, column=col_num).border = thin_border

            log_text.insert(tk.END, f"Process finished for column: {col}\n")
            log_text.update()

    log_text.insert(tk.END, "Fin.\n")
    log_text.update()


def start_process_threaded():
    threading.Thread(target=start_process).start()


root = tk.Tk()
root.title("Transcryptomic filter")
root.grid_columnconfigure(1, weight=1)
root.grid_rowconfigure(6, weight=1)

label_file1 = tk.Label(root, text="File go.xlsx:")
label_file1.grid(row=0, column=0, padx=10, pady=10, sticky="w")
entry_file1 = tk.Entry(root, width=50)
entry_file1.grid(row=0, column=1, padx=10, pady=10)
browse_file_button1 = tk.Button(root, text="...", command=lambda: browse_file(entry_file1))
browse_file_button1.grid(row=0, column=2, padx=10, pady=10)

label_file2 = tk.Label(root, text="File all_compare.xlsx:")
label_file2.grid(row=1, column=0, padx=10, pady=10, sticky="w")
entry_file2 = tk.Entry(root, width=50)
entry_file2.grid(row=1, column=1, padx=10, pady=10)
browse_file_button2 = tk.Button(root, text="...", command=lambda: browse_file(entry_file2))
browse_file_button2.grid(row=1, column=2, padx=10, pady=10)

label_function = tk.Label(root, text="Function: (ie. Chaperon)")
label_function.grid(row=2, column=0, padx=10, pady=10, sticky="w")
entry_function = tk.Entry(root, width=50)
entry_function.grid(row=2, column=1, padx=10, pady=10)

label_columns = tk.Label(root, text="Columns: (ie. HEK84_G50vsHEK)")
label_columns.grid(row=3, column=0, padx=10, pady=10, sticky="w")
entry_columns = tk.Entry(root, width=50)
entry_columns.grid(row=3, column=1, padx=10, pady=10)

label_dir = tk.Label(root, text="Saving dir:")
label_dir.grid(row=4, column=0, padx=10, pady=10, sticky="w")
entry_dir = tk.Entry(root, width=50)
entry_dir.grid(row=4, column=1, padx=10, pady=10)
browse_dir_button = tk.Button(root, text="...", command=lambda: browse_directory(entry_dir))
browse_dir_button.grid(row=4, column=2, padx=10, pady=10)

start_button = tk.Button(root, text="Start", command=start_process_threaded)
start_button.grid(row=5, column=0, columnspan=3, pady=20)

log_text = tk.Text(root, height=5, width=70)
log_text.grid(row=6, column=0, columnspan=3, padx=10, pady=10, sticky="nsew")

root.mainloop()
