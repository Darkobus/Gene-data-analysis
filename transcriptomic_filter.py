import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

"""
This script performs a targeted search for a specific phrase within a column in a gene database (Excel format). 
It then correlates these findings with your transcriptomics results to isolate only those entries of interest, 
specifically those with a pValue greater than 0.05. 
In the final step, it visually distinguishes the 'change' column: positive changes are highlighted in red, 
while negative changes are marked in blue.
"""

# Config data

# Path to files
base_go = 'gomice.xls.xlsx'
base_all_compared = 'all_comparemice.xls.xlsx'

# Desired phrase - Autophagy, Mitochondria etc
interested_function = 'chaperon'

# Column to compare
compare = ['R61vsWT_padj']


def generate_unique_filename(base_filename):
    """
    Generates a unique filename by appending a numerical suffix if the specified base filename already exists.
    :param base_filename: The base name of the file, including the path and extension.
    :return: A unique filename with a numerical suffix if necessary.
    """
    filename, extension = os.path.splitext(base_filename)
    counter = 1
    new_filename = base_filename

    while os.path.exists(new_filename):
        new_filename = f"{filename}({counter}){extension}"
        counter += 1

    return new_filename


def read_excel_file(file_path):
    try:
        df = pd.read_excel(file_path)
        print(f"Data imported form {file_path}.")
        return df
    except FileNotFoundError:
        print(f"Error: File {file_path} not found.")
        return None
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        return None


def main():
    print(f"Loading data from {base_go}...")
    df_go = read_excel_file(base_go)

    if df_go is not None:
        print("Filtering data form phrase...")
        filtered_go = df_go[df_go['go_term'].str.contains(interested_function, case=False)]
        print("Filtering done.")

        count_filtered = filtered_go.shape[0]
        print(f"Number of of '{interested_function}': {count_filtered}")

        print(f"Loading data from {base_all_compared}...")
        df_all_compared = pd.read_excel(base_all_compared, engine='openpyxl')
        print("Loaded succesfully.")

        print("Filtering and ordering ...")
        filtered_all_compared = df_all_compared[df_all_compared['gene_id'].isin(filtered_go['gene_id'])]

        print("pValue filtering...")
        condition = pd.Series(False, index=filtered_all_compared.index)
        for col in compare:
            condition |= (filtered_all_compared[col] < 0.05)

        count_condition = condition.shape[0]
        print(f"Number of results with padj < 0,05 '{condition}': {count_condition}")

        significant_genes = filtered_all_compared[condition]

        # Here's where we adjust the code
        # First, save the initial workbook
        initial_filename = 'Results.xlsx'
        significant_genes.to_excel(initial_filename, index=False)

        # Load the workbook to apply styles
        wb = load_workbook(initial_filename)
        ws = wb.active

        red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        blue_fill = PatternFill(start_color='FF0000FF', end_color='FF0000FF', fill_type='solid')

        for col_name in compare:
            try:
                col_index = [cell.value for cell in ws[1]].index(col_name)
            except ValueError:
                print(f"Column '{col_name}' not found in the Excel header. Skipping...")
                continue

            target_col_index = col_index - 2
            if target_col_index < 0:
                print(
                    f"Column '{col_name}' is too close to the start of the sheet to move two columns left. Skipping...")
                continue

            # Apply the fill to the target column
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=target_col_index + 1)
                if cell.value and cell.value < 0.05:
                    cell.fill = red_fill
                else:
                    cell.fill = blue_fill

        # Generate a unique filename for the styled workbook
        unique_filename = generate_unique_filename(initial_filename)

        # Save the styled workbook
        wb.save(unique_filename)
        print(f"Saving '{unique_filename}'... Praise beer!")


if __name__ == "__main__":
    main()
