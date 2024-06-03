Oczywiście, poniżej znajduje się zaktualizowany kod z komunikatami `print` przetłumaczonymi na język polski:

```python
import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from pathlib import Path

# Szybka konfiguracja

BASE_GO = 'path.xlsx'
BASE_ALL_COMPARED = 'path.xlsx'
INTERESTED_FUNCTION = 'chaperon'
COMPARE = ['HEK41vsHEK_padj', 'HEK41_G50vsHEK41_padj','HEK41_G50vsHEK_padj','HEK84vsHEK_padj','HEK84_G50vsHEK84_padj','HEK84_G50vsHEK_padj','HEK53vsHEK_padj', 'HEK453_G50vsHEK53_padj','HEK53_G50vsHEK_padj']

CHUNK_SIZE = 1000

def generate_unique_filename(base_filename):
    """
    Generuje unikalną nazwę pliku, dodając numeryczny sufiks, jeśli podana nazwa pliku już istnieje.
    :param base_filename: Podstawowa nazwa pliku, w tym ścieżka i rozszerzenie.
    :return: Unikalna nazwa pliku z numerycznym sufiksem, jeśli to konieczne.
    """
    filename = Path(base_filename)
    counter = 1  # Rozpocznij licznik dla sufiksu
    new_filename = filename  # Zainicjuj oryginalną nazwą

    # Pętla, aż znajdziesz unikalną nazwę pliku
    while new_filename.exists():
        new_filename = filename.with_name(f"{filename.stem}({counter}){filename.suffix}")
        counter += 1

    return new_filename

def process_chunk(chunk, expected_columns):
    """Przetwarzaj każdą część: zweryfikuj kolumny i wykonaj dalsze przetwarzanie."""
    if not set(expected_columns).issubset(chunk.columns):
        raise ValueError("Brakuje jednej lub więcej oczekiwanych kolumn w pliku Excel.")
    # Placeholder dla dalszego przetwarzania części
    print(chunk.head())  # Przykładowa akcja: wyświetl pierwsze kilka wierszy

def read_excel_file_in_chunks(file_path, chunk_size):
    try:
        # Próba zainicjowania czytnika z podziałem na części
        reader = pd.read_excel(file_path, chunksize=chunk_size)
    except FileNotFoundError:
        print(f"Błąd: Plik {file_path} nie został znaleziony.")
        return
    except Exception as e:
        print(f"Wystąpił nieoczekiwany błąd: {e}")
        return

    expected_columns = ['gene_id', 'go_term']
    try:
        for chunk in reader:
            process_chunk(chunk, expected_columns)
    except ValueError as e:
        print(e)
        return

def analyze_column(df_go, df_all_compared, col):
    print(f"Analiza kolumny: {col}")
    filtered_go = df_go[df_go['go_term'].str.contains(INTERESTED_FUNCTION, case=False)]

    count_filtered = filtered_go.shape[0]
    print(f"Liczba komórek zawierających frazę '{INTERESTED_FUNCTION}': {count_filtered}")

    filtered_all_compared = df_all_compared[df_all_compared['gene_id'].isin(filtered_go['gene_id'])]

    condition = filtered_all_compared[col] < 0.05
    filtered_df = filtered_all_compared[condition]

    count_condition = filtered_df.shape[0]
    print(f"Liczba wyników spełniających warunek: {count_condition}")

    return filtered_df

def color_columns(workbook, sheet_name, compare_col_index):
    ws = workbook[sheet_name]
    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    blue_fill = PatternFill(start_color='FF0000FF', end_color='FF0000FF', fill_type='solid')

    target_col_index = compare_col_index - 2
    for row in range(2, ws.max_row + 1):
        target_cell = ws.cell(row=row, column=target_col_index)
        compare_cell = ws.cell(row=row, column=compare_col_index)

        # Upewnij się, że komórka porównawcza ma wartość
        if compare_cell.value is not None:
            if compare_cell.value > 0:
                target_cell.fill = red_fill
            elif compare_cell.value < 0:
                target_cell.fill = blue_fill

def main():
    print(f"Ładowanie bazy danych {BASE_GO}...")
    cols_to_read = ['gene_id', 'go_term']
    df_go = pd.read_excel(BASE_GO, usecols=cols_to_read)

    print(f"Ładowanie danych z {BASE_ALL_COMPARED}...")
    df_all_compared = pd.read_excel(BASE_ALL_COMPARED, engine='openpyxl')
    print("Ładowanie danych zakończone.")

    unique_wyniki = generate_unique_filename("Wyniki.xlsx")
    with pd.ExcelWriter(unique_wyniki, engine='openpyxl') as writer:
        for col in COMPARE:
            filtered_df = analyze_column(df_go, df_all_compared, col)
            filtered_df.to_excel(writer, sheet_name=col, index=False)
    print("Wszystkie dane zostały zapisane na oddzielnych arkuszach.")

    # Załaduj skoroszyt, aby zastosować kolorowanie
    wb = load_workbook(unique_wyniki)
    for col in COMPARE:
        ws = wb[col]
        compare_col_index = ws[1].index([cell for cell in ws[1] if cell.value == col][0]) + 1
        color_columns(wb, col, compare_col_index)

    unique_filename = generate_unique_filename(f'{INTERESTED_FUNCTION}.xlsx')
    print(f"Zapisywanie pokolorowanego pliku '{unique_filename}'...")
    wb.save(unique_filename)
    print("Zapisywanie zakończone. Na zdrowie!")

if __name__ == "__main__":
    main()
```
